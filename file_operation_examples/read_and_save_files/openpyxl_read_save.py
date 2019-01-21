# -*- coding: utf-8 -*-
"""
Created on Mon Jan 21 17:33:51 2019

@author: lankuohsing
"""

# coding:utf-8
from openpyxl import load_workbook
import openpyxl

# In[]
wb = load_workbook("openpyxl_test1.xlsx")#生成一个已存在的wookbook对象
# In[]
"""获取工作表--Sheet"""
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('Sheet1')
# 获得sheet名

# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.get_sheet_by_name('Sheet1')
# In[]
"""获取单元格"""
# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['B4']
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型
"""
Python3.6新增了一种f-字符串格式化
格式化的字符串文字前缀为’f’和接受的格式字符串相似str.format()。它们包含由花括号包围的替换区域。替换字段是表达式，在运行时进行评估，然后使用format()协议进行格式化。
formatted string literals, 以 f 开头，包含的{}表达式在程序运行时会被表达式的值代替。
"""
# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=4, column=2)
print(b4_too.value)
# In[]
"""获得最大行和最大列"""
# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)
# In[]
"""
获取行和列
sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
sheet.columns类似，不过里面是每个tuple是每一列的单元格。
"""
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)
# In[]
list1=list(sheet.rows)
print(list1[0][0].value)
# In[]
sheet["A1"]=100
sheet.cell(3,3,1000)
# In[]
wb.save('openpyxl_test2.xlsx')
