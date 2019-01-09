from openpyxl import load_workbook
wb2 = load_workbook('balances.xlsx')
print(wb2.sheetnames)
sheet_ranges = wb2['Mysheet1']
print(sheet_ranges['A1'].value)
