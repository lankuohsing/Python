from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet1") # insert at the end (default)
# or
ws2 = wb.create_sheet("Mysheet2", 0) # insert at first position
ws.title = "New Title"
ws3 = wb["New Title"]
ws1["A1"]="A1"
ws1["B1"]="B1"
ws1.append(["A2","B2"])
ws2["A2"]="A2"
ws['A4'] = 4
ws['C9'] = 'hello world'
wb.save('balances.xlsx')
